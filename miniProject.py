import time
import selenium.webdriver
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.chrome.options import Options

from colorama import Fore, Back, Style

import string
import pandas as pd
import openpyxl as op
import os
import xlsxwriter

class scraping:
	# Creating driver and other parameters to open an automated web browser.. init is used for self declaration
	def __init__(self):
		self.driver=webdriver.Chrome(executable_path ='/opt/chromedriver')
	        #self.options = Options()
	        #self.options.add_argument('--headless')
	        #self.options.add_argument('--disable-gpu')  # Last I checked this was necessary.
	        #self.driver = webdriver.Chrome('/opt/chromedriver', chrome_options=self.options)
		self.marks={}
		self.cod={}
		self.count=0
		self.j=1
		self.list1=[]
		self.flist=[]
		self.fail=[]
		self.writename='Result.xlsx'
		self.filename='Result.xlsx'
		self.flag=1
		
	# Range method.. self is used for referencing the object calling the method.. y1 and y2 denotes the range
	def usn_by_range(self,mini,maxi):
		for x in range(mini,maxi+1):
			x= '{0:03}'.format(x)
			usn='1ox15is' + x
			self.scrape(usn)
		self.write_file(self.writename)
	
			
	# Fetching from file... filename is the name of file which user is giving... r denotes the read mode in which file is opened... scrape(usn) is calling scraping method by passing a parameter 'usn' 
	
	def usn_by_file(self,filename):
		try:
			with open(filename,'r') as f:
				for n in f:
					if n=='\n':
						continue
					usn=str(n)[:-1]
					self.scrape(usn)
				self.write_file(self.writename)
		
		except Exception as e:
			print(Fore.RED + "FILE NOT FOUND.... Please try again..");
		
	# Manually asking user to input the usn
	def usn_manual(self):
		# semester
		sems=raw_input("Enter the sem: ")
		
		# Region code
		while True:
			reg_code=raw_input("Please Enter the region code: \n")
			if reg_code <= '4' :
				break
			else:
				print "Wrong region code.. valid values are (1,2,3,4) \nTry again"
			
		# College code
		clg_code=raw_input("Please Enter the college code: \n").lower()
		
		# year or batch
		bat_code=raw_input("Please Enter year: \n").lower()
		
		# branch
		branch_code=raw_input("Please Enter branch code: \n").lower()
		
		print("Enter the range of usn's: ")
		mini = int(raw_input("Minimum: "))
		maxi = int(raw_input("Maximum: "))
		usn_build= reg_code + clg_code + bat_code + branch_code		

		if len(branch_code) == 3:
			for x in range(mini,maxi+1):	
				x= '{0:02}'.format(x)
				usn= usn_build + x
				self.scrape(usn)

		else:
			for x in range(mini,maxi+1):	
				x= '{0:03}'.format(x)
				usn= usn_build + x
				self.scrape(usn)
				
		self.write_file(self.writename)	
			
	def single_usn(self):
		usn = raw_input("Enter the usn of student: ")
		self.scrape(usn)
		if self.flag==1:
			self.write_file(self.writename)	
			
				
	def scrape(self,usn):
		try:
			# Browse website
			self.driver.get("http://results.vtu.ac.in/vitaviresultcbcs/index.php")
		
			# Searching element USN
			element = self.driver.find_element_by_xpath('/html/body/div[2]/div[1]/div[2]/div/div[2]/form/div/div[2]/div[1]/div/input')
			
			# Sending usn value
			element.send_keys(usn)
		
			# Click on Submit button
			self.driver.find_element_by_xpath('//*[@id="submit"]').click()
		
			try:
				# To accept the alert message
				if selenium.webdriver.support.expected_conditions.alert_is_present():
					alert = self.driver.switch_to_alert()
					print("\n"+Style.BRIGHT + Fore.GREEN + usn + " "+ Fore.RED + "\t" +"USN NOT VALID")
					self.flag=0
					alert.accept()
					
			except Exception as e:
				#semester
				sem=self.driver.find_element_by_xpath('//*[@id="dataPrint"]/div[2]/div/div/div[2]/div[1]/div/div/div[2]/div/div/div[1]').text
			
				# Formatting to remove spaces and special character like ':'
				if(str(sem).translate(None,string.punctuation+string.whitespace).lower()!= "semester5"):
					print("\n"+Style.BRIGHT + Fore.GREEN + usn + " "+ Fore.RED + "\t" +"YEAR BACK STUDENT")
					pass
			
				else:
				
					# Fetching USN
					usnn=self.driver.find_element_by_xpath('//*[@id="dataPrint"]/div[2]/div/div/div[2]/div[1]/div/div/div[1]/div/table/tbody/tr[1]/td[2]').text
					
					# Formatting USN to remove spaces and special characters like ':'
					usnn=str(usnn).translate(None,string.punctuation+string.whitespace)
	
					# Feeding into list and dictionary so that it can be used in dataframe at 0
					self.marks["usn"]=usnn
					self.list1.append(usnn)
		
					# Fetching Name
					name=self.driver.find_element_by_xpath('//*[@id="dataPrint"]/div[2]/div/div/div[2]/div[1]/div/div/div[1]/div/table/tbody/tr[2]/td[2]').text
					
					# Formatting Name to remove spaces and special characters like ':'
					name=str(name).translate(None,string.punctuation)
	
					# Feeding into list and dictionary so that it can be used in dataframe 1
					self.marks["name"]=name
					self.list1.append(name)
					
					# fetching subject codes through web scraping
					for i in range(2,10):
				
						# Storing subject codes in an array 'cod'
						self.cod[i]=str(self.driver.find_element_by_xpath('//*[@id="dataPrint"]/div[2]/div/div/div[2]/div[1]/div/div/div[2]/div/div/div[2]/div/div['+str(i)+ ']/div[1]').text)
						# Also creating a sub dictionary of subject codes
						self.marks[self.cod[i]]=[]
								
				
					# Printing usn and name on terminal and Format of subjects
					print("\n"+Style.BRIGHT + Fore.GREEN + usnn + " "+ name)
					print(Fore.MAGENTA +Style.DIM + "\t\t\t\tSubject\t\tInt\tExt\tTotal\tResult"+Style.RESET_ALL)
					
					
					# Initialising marks['index'] to store some parameters	
					self.marks["result"]='Pass'
					self.marks["total"]=0
					self.marks["sub_fail"]=[]
				
					# Fetching the marks for each subject and storing them in respective dictionaries
					for i in range(2,10):
						for k in range(3,7):
							self.marks[str(self.driver.find_element_by_xpath('//*[@id="dataPrint"]/div[2]/div/div/div[2]/div[1]/div/div/div[2]/div/div/div[2]/div/div['+str(i)+']/div[1]').text)].append(str(self.driver.find_element_by_xpath('//*[@id="dataPrint"]/div[2]/div/div/div[2]/div[1]/div/div/div[2]/div/div/div[2]/div/div['+str(i)+']/div['+str(k)+']').text))
					
						# Printing marks details on terminal							
						print(Fore.WHITE + Style.BRIGHT +"\t\t\t\t"+self.cod[i] +"\t\t"+self.marks[self.cod[i]][0]+"\t"+self.marks[self.cod[i]][1] +"\t"+self.marks[self.cod[i]][2]+"\t"+self.marks[self.cod[i]][3])
				
						
						
						# Appending to the list now
							
						self.list1.append(self.marks[self.cod[i]][0])
						self.list1.append(self.marks[self.cod[i]][1])
						self.list1.append(self.marks[self.cod[i]][2])
			
						# Calculating total marks
						self.marks["total"]=self.marks["total"]+int(self.marks[self.cod[i]][2])
						# Checking whether failed or not
						if(self.marks[self.cod[i]][3]!='P'):
							#if count==0:
							self.marks["result"]='Fail'
					
							self.count = self.count+1	
							self.fail.append(self.cod[i])
							self.marks['sub_fail'].append(self.cod[i])				
						else:
							pass	
					# Calculating Percentage
					self.marks["percent"]=self.marks["total"]/8
					
					#Writing no. of subjects failed
					self.marks["failed"]=self.count
					
					
					if(self.count>0):
						print(Fore.MAGENTA + "\t\tResult: " + Fore.RED + "Fail")
						print(Fore.MAGENTA + Style.BRIGHT + "\t\tPercent: "+ Fore.RED +str(self.marks["percent"])+ "%")
						print(Fore.MAGENTA +Style.DIM + "\n\t\tFailed in : "+ Fore.RED + str(self.marks['sub_fail'])+Style.RESET_ALL)
					else:
						
						print(Fore.MAGENTA + "\t\tResult: " + Fore.CYAN + "Pass")
						print(Fore.MAGENTA + Style.BRIGHT + "\t\tPercent: "+ Fore.CYAN +str(self.marks["percent"])+ "%")
						
					
					# Appending all necessary information to list
					self.list1.append(self.marks["result"])
					self.list1.append(self.marks["total"])
					self.list1.append(self.marks["percent"])
					self.list1.append(self.count)
					self.list1.append(self.marks["sub_fail"])			
					
					# Appending to final list
					self.flist.append(self.list1)
					self.count=0
					# Clearing result
					self.list1=[]
				
				
			
				pass
			
		except Exception as e:
			print("Exception"+str(e))
			self.write_file(self.writename)

			
	def open_file(self):
		file_name=raw_input("Enter the name of file with proper extension: ")
		cmd= "/usr/bin/xdg-open /home/ayush/miniProject/" + file_name
		try:
			os.system(cmd)
		except:
			print(Fore.RED + "FILE NOT FOUND.... Please try again..");	

	def search(self):
		file_name = raw_input("Enter the name of file with proper extension: ")
		keyword = raw_input("Enter the string to search: ").upper()
		try:
		
			wb=op.load_workbook('/home/ayush/miniProject/'+file_name)
			sheet= wb['Sheet1']
			ws=wb.active
			sflag=0
			try:
				for x in range(1,ws.max_row+1):
					
					if ws.cell(x,1).value==keyword:
						sflag=1
						print("\n")
						for y in range(1,ws.max_column+1):
							if(ws.cell(1,y).value==None):
	                		         		print " |", 
	                				else:
	                		 		       print Fore.MAGENTA + ws.cell(1,y).value + " | ",
	                			print("")
	                	 
	                			for y in range(1,ws.max_column+1):
	                		     		print Fore.WHITE + Style.BRIGHT + str(ws.cell(x,y).value) + " | ",
	                		     	break
	                     		else:
	                     			continue
	                     		Style.RESET_ALL
	    		except Exception as e:
	    			print(Fore.RED + str(e))
	      		    	pass
		
			if sflag==0:
				print(Fore.RED + "Record not found"+Style.RESET_ALL)
		except Exception as p:
			print(Fore.RED + "File not found" + Style.RESET_ALL)

					
	def write_file(self,filename):
		try:
		# Creating a dataFrame
			data = pd.DataFrame(self.flist, columns=('USN_USN', 'Name_Name', self.cod[2]+'_int', self.cod[2]+'_ext', self.cod[2]+'_total', self.cod[3]+'_int', self.cod[3]+'_ext', self.cod[3]+'_total', self.cod[4]+'_int', self.cod[4]+'_ext', self.cod[4]+'_total', self.cod[5]+'_int', self.cod[5]+'_ext', self.cod[5]+'_total', self.cod[6]+'_int', self.cod[6]+'_ext', self.cod[6]+'_total',  self.cod[7]+'_int', self.cod[7]+'_ext', self.cod[7]+'_total', self.cod[8]+'_int', self.cod[8]+'_ext', self.cod[8]+'_total', self.cod[9]+'_int', self.cod[9]+'_ext', self.cod[9]+'_total','result_','total_','percent_','no. of subject failed_','subject failed_'))
	
		
			data.columns = data.columns.str.split('_', expand=True)
			data.columns = data.columns.swaplevel(1,0)
		
		# Writing into file
			writer = pd.ExcelWriter(filename, engine='xlsxwriter')
			data.to_excel(writer,sheet_name='Sheet1')
			workbook=writer.book
			merge_format = workbook.add_format({
	     	    'bold': 1,
	    	     'border': 1,
	    	     'align': 'center',
	    	     'valign': 'vcenter',
	    	     'fg_color': 'yellow'})
			worksheet=writer.sheets['Sheet1']
			worksheet.merge_range('A1:A2',' ',merge_format)
		    	worksheet.merge_range('B1:B2','USN',merge_format)
			worksheet.freeze_panes(1, 0)
			worksheet.freeze_panes(2, 0)
			writer.save()
		
		# Opening a file and editing it
			wb=op.load_workbook(filename)
			ws=wb.active
			ws.delete_rows(3)
			ws.delete_cols(1)
			wb.save(filename)
			#self.flist=[]
		except Exception as e:
			print(Fore.RED + "1 "+str(e))
			print(e)
					
	def read_file(self):
		file_name = raw_input("Enter the name of file with proper extension: ")
		try:
		
			df=pd.read_excel(file_name)
			print df
		except:
			print(Fore.RED + Style.BRIGHT + "File not found"+Style.RESET_ALL)

	def delete_entry(self):
		file_name = raw_input("Enter the name of file with proper extension: ")
		keyword = raw_input("Enter the usn you want to delete: ").upper()
		try:
			wb=op.load_workbook(file_name)
			ws=wb.active
			sflag=0
			for x in range(1,ws.max_row+1):
				try:
					if ws.cell(x,1).value==keyword:
						print(Fore.GREEN + "\nRecord found...")
						print(Fore.RED + Style.BRIGHT + "Deleting it.."+ Style.RESET_ALL)
						sflag=1
						ws.delete_rows(x)
						
						wb.save(file_name)
						
						break
						
        	             		else:
        	             			continue
        	             			
        	        	except Exception as e:
      			      		pass
      			
      			if sflag==0:
      				print(Fore.RED + Style.BRIGHT + "Record not found..."+ Style.RESET_ALL) 
      		
      		except Exception as p:
      			print(Fore.RED + "File not found"+Style.RESET_ALL)
      					
      	def delete_allentry(self):
      		file_name = raw_input("Enter the name of file with proper extension: ")		
      		try:
	      		wb=op.load_workbook(file_name)
			ws=wb.active
			self.flist=[]
			self.write_file(file_name)
		except Exception as e :
			print(Fore.RED + "File not found"+Style.RESET_ALL)
				
	def modify_entry(self):
		file_name = raw_input("Enter the name of file with proper extension: ")
		keyword = raw_input("Enter the usn you want to modify: ").upper()
		try:
			wb=op.load_workbook(file_name)
			sheet= wb.get_sheet_by_name('Sheet1')
			ws=wb.active
			sflag=0
			for x in range(1,ws.max_row+1):
				try:
					if ws.cell(x,1).value==keyword:
						sflag=1
						for y in range(1,ws.max_column+1):
							if(ws.cell(1,y).value==None):
        	        		         		pass 
        	        				else:
        	        		 		       print Fore.WHITE+Style.BRIGHT+"\nContent of "+ws.cell(1,y).value+" :",
        	        		 		       print(ws.cell(x,y).value + Style.RESET_ALL)
        	        		 		       modify=raw_input("Modify it as: (press "+Fore.BLUE+Style.BRIGHT +"'enter'"+Style.RESET_ALL +" to skip or type "+ Fore.BLUE + Style.BRIGHT +"'done'"+ Style.RESET_ALL+" to finish): ").upper()
        	        		 		       
        	        		 		       if modify=='' :
        	        		 		       		pass
        	        		 		       elif modify=='DONE':
        	        		 		       		break
        	        		 		       
        	        		 		       else:
        	        		 		       		ws.cell(x,y).value=modify		
                		 		      
						wb.save(file_name)
						print(Fore.GREEN + Style.BRIGHT+ "Modification done"+ Style.RESET_ALL)
						break
					
        	             		else:
        	             			continue
                     			
        	        	except Exception as e:
        	        		print(Fore.RED +"2"+ str(e))
      			      		pass
      		
      			if sflag==0:
      				print "Record not found"  
      			
      		except Exception as p:
      			print(Fore.RED + "File not found" + Style.RESET_ALL)	
      				    		
def main():
    Style.RESET_ALL
    print("\n\n"+ Fore.YELLOW + Style.BRIGHT+ " :-:-:-:-:-:-:-:-:-:-:-:-:-:-:->>"+Fore.CYAN+" Welcome to WEB SCRAPING OF RESULT with FILES Project"+Fore.YELLOW+" <<-:-:-:-:-:-:-:-:-:-:-:-:-:-:-: ")
    s1 = scraping()
    mini=1
    maxi=5
    while 1 :

        print(Style.BRIGHT + Fore.CYAN + "\n" + "\tWhat would you like to do?")
        print(Style.RESET_ALL)
        print(Fore.YELLOW + "\t1) Web Scraping of Result")
        print("\t2) Operations in File")
        print("\t3) Exit")

        print(Fore.CYAN + Style.BRIGHT+"\nEnter your choice: "+Style.RESET_ALL)


        choice = raw_input()

        if choice == '1':
	    
	    while 1:
	            print(Style.BRIGHT + Fore.CYAN + "\n\tEXCEUTING WEB SCRAPING, Take usn from: " + Style.RESET_ALL)
        	    print(Fore.YELLOW + "\t1) Set of USN's")
        	    print("\t2) File")
        	    print("\t3) Manually")
        	    print("\t4) Particular USN")
        	    print("\t5) Go back")
        	    print("\t6) Exit")
	
        	    scrape_choice = raw_input(Fore.CYAN + Style.BRIGHT+"\nEnter your choice: "+Style.RESET_ALL)
	
	
        	    if scrape_choice == '1':
        	        mini=int(raw_input("\nEnter the minimum: "))
        	        maxi=int(raw_input("\nEnter the maximum: "))
        	        s1.usn_by_range(mini,maxi)
	
        	    elif scrape_choice == '2':
        	        file_name=raw_input("\nEnter the name of file with proper extension: ")
        	        s1.usn_by_file(file_name)
	
        	    elif scrape_choice == '3':
        	        s1.usn_manual()
	
        	    elif scrape_choice == '4':
        	        s1.single_usn()
	
        	    elif scrape_choice == '5':
        	        pass
        	        break;
	
        	    elif scrape_choice == '6':
        	        exit(0)
	
        	    else:
        	        print(Fore.RED + Style.BRIGHT + "\t\t\t\tInvalid choice... .................Please try again............"+ Style.RESET_ALL)



        elif choice == '2':
	
	    while 1:
	    
            	print(Style.BRIGHT + Fore.CYAN + "\n\tEXCEUTING OPERATIONS in FILE: " + Style.RESET_ALL)
            	print(Fore.YELLOW+"\t1) Open a file in default application")
            	print("\t2) Search within a file")
            	print("\t3) Reading a complete file")
	    	print("\t4) Modify a particular record in a file")
	    	print("\t5) Delete a particular record")
	    	print("\t6) Delete all records from file")
            	print("\t7) Go back")
            	print("\t8) Exit")

            	file_choice = raw_input(Fore.CYAN + Style.BRIGHT+"\nEnter your choice: "+Style.RESET_ALL)

            	if file_choice == '1':
            	    s1.open_file()

            	elif file_choice == '2':
            	    s1.search()

            	elif file_choice == '3':
            	    s1.read_file()

 	    	elif file_choice == '4':
	    	    s1.modify_entry()
 
            	elif file_choice == '5':
	    	    s1.delete_entry()
            
            	elif file_choice == '6':
	    	    s1.delete_allentry()
           
            	elif file_choice == '7':
		    break

	        elif file_choice == '8':
	            exit(0)
	    
	    
	    
           	else:
                    print(Fore.RED + Style.BRIGHT + "\t\t\t\tInvalid choice... .................Please try again............"+Style.RESET_ALL)

        elif choice == '3':
            exit(0)

        else:
            print(Fore.RED + Style.BRIGHT+"\t\t\t\tInvalid choice... .................Please try again............"+Style.RESET_ALL)

main()
Style.RESET_ALL
